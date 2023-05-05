import calendar
import datetime
import logging
import random
from typing import List, Union

import openpyxl
from django.forms import forms
from django.utils import timezone
from openpyxl.cell import MergedCell, Cell

from core.models import UploadedFile, StatDataItem


logger = logging.getLogger(__name__)


class XLSParser:
    """
    Main class for Excel files parsing
    """
    def __init__(self, file):
        self.workbook = openpyxl.load_workbook(file)
        self.worksheet = self.workbook.worksheets[0]
        self.data_rows_start_index = None
        self.field_names = self.__parse_header()
        self.table_width = len(self.field_names)

    def __merged_cell_value(self, cell: MergedCell):
        """
        Searches value of provided merged cell in start cell of its merged range
        :param cell: MergedCell object
        :return: value of start cell or empty string
        """
        possible_ranges = [r for r in self.worksheet.merged_cells.ranges if cell.coordinate in r]
        if not possible_ranges:
            return ''
        else:
            merged_range = possible_ranges[0]
            return merged_range.start_cell.value

    def __get_cell_value(self, cell: Union[Cell, MergedCell]):
        """
        Processor that uses direct call to cell value or searches value for merged cell
        :param cell: Cell or MergedCell object
        :return: value or empty string
        """
        if isinstance(cell, MergedCell):
            return self.__merged_cell_value(cell)
        return cell.value

    @staticmethod
    def __cut_off_tail(tailed_list: List[list]):
        """
        Cuts off empty lists from the end of the provided list
        Needed for correct counting the width of the data table
        :param tailed_list: list of lists
        :return: untailed list
        """
        while not tailed_list[-1]:
            tailed_list.pop(-1)
        return tailed_list

    def __parse_header(self):
        """
        Parses the header of the data table and automatically
        generates the list of field names to use in database scheming
        :return: list of strings
        """
        header_merged_cells = list(filter(lambda x: x.min_row == 1,
                                          self.worksheet.merged_cells.ranges))
        header_width = max([i.max_row for i in header_merged_cells])
        self.data_rows_start_index = header_width + 1
        field_names = [list() for _ in range(self.worksheet.max_row)]
        for row in self.worksheet.iter_rows(min_col=1, min_row=1, max_row=header_width):
            for t, cell in enumerate(row):
                value = self.__get_cell_value(cell)
                if value is not None and (not field_names[t] or field_names[t][-1] != value):
                    field_names[t].append(str(value).lower())
        return ['_'.join(i) for i in self.__cut_off_tail(field_names)]

    @property
    def data(self):
        """
        Converts data from Openpyxl worksheet cells to the list of dicts
        :return: list of dicts
        """
        result = list()
        for row in self.worksheet.iter_rows(min_col=1, max_col=self.table_width, min_row=self.data_rows_start_index):
            row_dict = dict()
            for t, cell in enumerate(row):
                row_dict[self.field_names[t]] = cell.value
            result.append(row_dict)
        return result


class RandomDatePicker:
    """
    Random in-month date generator.
    Needed for addition random dates to the database (according to the condition of the task)
    """

    def __init__(self, year: int = None, month: int = None):
        if year is None:
            current_year = timezone.now().year
            self.year = random.randint(current_year - 5, current_year)
        else:
            self.year = self.__validate_year(year)

        if month is None:
            self.month = random.randint(1, 12)
        else:
            self.month = self.__validate_year(year)

        self.dates = [date for date in self.__dates_list() if date.month == self.month]

    @staticmethod
    def __validate_year(year: int):
        """
        Validates year if provided. Raises an exception if value does not meet the condition
        :param year: value to validate
        :return: int or throws an exception
        """
        if not isinstance(year, int) or not 0 < year <= 9999:
            raise ValueError(f'Year must be integer from 1 to 9999. Provided value is {repr(year)}')
        return year

    @staticmethod
    def __validate_month(month: int):
        """
        Validates month if provided. Raises an exception if value does not meet the condition
        :param month: value to validate
        :return: int or throws an exception
        """
        if not isinstance(month, int) or not 0 < month <= 12:
            raise ValueError(f'Year must be integer from 1 to 12. Provided value is {repr(month)}')
        return month

    def __dates_list(self):
        """
        Generates list of possible dates in the needed month
        :return: list of dates
        """
        dates_generator = calendar.Calendar().itermonthdates(self.year, self.month)
        return [date for date in dates_generator if date.month == self.month]

    def pick_date(self):
        """
        Random date picker
        :return: random date of generated list
        """
        return random.choice(self.dates)


class XLSUploadForm(forms.Form):
    """
    Django form for parsing the file provided
    """
    file = forms.FileField(label='Upload Excel file')

    def __init__(self, *args, **kwargs):
        super(XLSUploadForm, self).__init__(*args, **kwargs)
        self.file_model = UploadedFile
        self.data_model = StatDataItem
        self.worksheet = None
        self.unnecessary_fields = ['primary_id', 'created_at', 'file', 'date']
        self.allowed_fields = [f.name for f in self.data_model._meta.fields if f.name not in self.unnecessary_fields]

    def add_header_error(self, message, error_fields):
        """
        Adds a header parsing error to the form errors
        :param message: explanation about the error
        :param error_fields: incorrect field names
        """
        if error_fields:
            error_msg = '{}: {}'.format(
                message,
                ', '.join(error_fields)
            )
            self.add_error('file', error_msg)

    def check_header(self):
        """
        makes human-readable error messages if some field names are missing or incorrect
        """
        self.add_header_error('Following names are missing',
                              [i for i in self.allowed_fields if i not in self.worksheet.field_names])
        self.add_header_error('Following names are not allowed',
                              [i for i in self.worksheet.field_names if i not in self.allowed_fields])

    def clean(self):
        """
        Extension for Django's default form cleaning method.
        Checks provided header in advance
        """
        file = super(XLSUploadForm, self).clean().get('file')
        try:
            self.worksheet = XLSParser(file)
            self.check_header()
        except Exception as e:
            logger.error(e)
            self.add_error('file', 'Incompatible file!')

    def preprocess_data(self):
        """
        Preprocessor for data in provided file
        :return: file model object and data prepared for saving in DB
        """
        file = self.file_model.objects.create(file_name=self.cleaned_data['file'].name)
        date_picker = RandomDatePicker()
        result = list()
        for item in self.worksheet.data:
            result.append(self.data_model(file=file, date=date_picker.pick_date(), **item))
        return file, result

    def save(self):
        """
        Saves data from file to the DB. In case of an exception deletes file object
        :return: file model object
        """
        file, data = self.preprocess_data()
        try:
            self.data_model.objects.bulk_create(data)
            return file
        except Exception as e:
            logger.error(e)
            file.delete()
            raise Exception('Something is wrong with provided data')
